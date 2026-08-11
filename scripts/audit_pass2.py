"""Pass 2 — everything the first render audit captured but never compared,
plus the charts and CSV exports it never captured at all.

Covers: ov-zero, ov-long, di-scatter, di-rank, bo-table, bo-dbar, ph-table,
sg-gap chart, tf-table, tf-delta, and all six CSV exports in full.
Independent recomputation from the original .xlsx, as before.
"""
import openpyxl, json, os, sys
from decimal import Decimal, ROUND_HALF_UP
R = os.path.join(os.path.dirname(__file__), '..')
YEARS=[2018,2019,2022,2023,2024,2025,2026]; MODERN=[2023,2024,2025,2026]
CATS=['All Students','SWD','Not SWD','Econ Disadv','Not Econ Disadv','Current ELL','Ever ELL',
      'Never ELL','Asian','Black','Hispanic','Multi-Racial','Native American','White',
      'Female','Male','Neither Female nor Male']
BOROS=['Bronx','Brooklyn','Manhattan','Queens','Staten Island']
BKEY={'BRONX':'Bronx','BROOKLYN':'Brooklyn','MANHATTAN':'Manhattan','QUEENS':'Queens','STATEN ISLAND':'Staten Island'}
DISTRICTS=['%02d'%d for d in range(1,33)]
BANDS={'all':['All Grades'],'35':['3','4','5'],'68':['6','7','8'],
       'g3':['3'],'g4':['4'],'g5':['5'],'g6':['6'],'g7':['7'],'g8':['8']}
BAND_LABEL={'all':'All grades (3–8)','35':'Grades 3–5 (elementary)','68':'Grades 6–8 (middle)'}
BORO_OF={}
for d in range(1,7): BORO_OF['%02d'%d]='Manhattan'
for d in range(7,13): BORO_OF['%02d'%d]='Bronx'
for d in list(range(13,24))+[32]: BORO_OF['%02d'%d]='Brooklyn'
for d in range(24,31): BORO_OF['%02d'%d]='Queens'
BORO_OF['31']='Staten Island'
PHASE={'elem1':[5,11,12,14,16,19,20,21,22,23,25,26,29,30,32],
       'elem2':[1,2,3,4,6,7,8,9,10,13,15,17,18,24,27,28,31],
       'ms1':[1,3,7,9,11,12,13,19],'ms2':[2,4,8,10,14,18,20,26,29,32]}

def load(path, idcol):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True); C={}
    for sn in wb.sheetnames:
        if not sn.startswith('ELA'): continue
        ws=wb[sn]; it=ws.iter_rows(values_only=True); hdr=list(next(it)); H={h:i for i,h in enumerate(hdr)}
        for r in it:
            if r is None or r[0] is None: continue
            cat=r[H['Category']]
            if cat not in CATS: continue
            v=[r[H[c]] for c in ('Number Tested','Mean Scale Score','# Level 1','# Level 2','# Level 3','# Level 4')]
            if any(x=='s' for x in v): continue
            gid = BKEY[r[H[idcol]]] if idcol=='Borough' else (r[H[idcol]] if idcol else 'city')
            C[(gid,str(r[H['Grade']]),r[H['Year']],cat)]=v
    return C
C={'city':load(os.path.join(R,'data','citywide-ela-results-public.xlsx'),None),
   'boro':load(os.path.join(R,'data','borough-ela-results-public.xlsx'),'Borough'),
   'dist':load(os.path.join(R,'data','district-ela-results-public.xlsx'),'District')}

def A(lvl,gids,bk,year,cat):
    n=c1=c2=c3=c4=0; wm=0.0; found=False
    for gid in gids:
        for g in BANDS[bk]:
            v=C[lvl].get((gid,g,year,CATS[cat] if isinstance(cat,int) else cat))
            if v is None: continue
            found=True; nt,ms,a,b,c,d=v
            n+=nt;c1+=a;c2+=b;c3+=c;c4+=d;wm+=ms*nt
    if not found or n==0: return None
    return dict(n=n,l1=c1/n*100,l2=c2/n*100,l3=c3/n*100,l4=c4/n*100,
                prof=(c3+c4)/n*100,mean=wm/n)
MET={'prof':'prof','l1':'l1','l4':'l4','mean':'mean'}
GOOD={'prof':1,'l1':-1,'l4':1,'mean':1}

def _r(v,nd):
    return str(Decimal(repr(float(v))).quantize(Decimal('1.'+'0'*nd if nd else '1'),
                                                rounding=ROUND_HALF_UP))
def f1(v): return '—' if v is None else _r(v,1)
def f2(v): return '' if v is None else _r(v,2)
def ppf(v):
    if v is None: return '—'
    if abs(v)<0.05: return '0.0'
    return ('+' if v>0 else '−')+_r(abs(v),1)
def numf(v): return f'{v:,}'

bad=[]; checked=0
def cmp(tag,got,exp,tol=None):
    global checked; checked+=1
    if tol is not None:
        if got is None and exp is None: return
        if got is None or exp is None: bad.append((tag,got,exp)); return
        if abs(got-exp)>tol: bad.append((tag,got,exp))
    elif str(got)!=str(exp): bad.append((tag,got,exp))

P2=json.load(open(os.path.join(R,'audit/p2_dump.json')))
for e in P2:
    p=e['p']
    if p=='ov2':
        zy,bk,dim,cat=e['st']; cat=int(cat)
        # diverging chart: L1/L2 plotted negative, L3/L4 positive, by grade
        for gi,g in enumerate(['g3','g4','g5','g6','g7','g8']):
            a=A('city',['city'],g,zy,cat)
            cmp(f'ov-zero L1 {e["st"]} {g}', e['zero'][0]['v'][gi], -a['l1'] if a else None, 1e-5)
            cmp(f'ov-zero L2 {e["st"]} {g}', e['zero'][1]['v'][gi], -a['l2'] if a else None, 1e-5)
            cmp(f'ov-zero L3 {e["st"]} {g}', e['zero'][2]['v'][gi],  a['l3'] if a else None, 1e-5)
            cmp(f'ov-zero L4 {e["st"]} {g}', e['zero'][3]['v'][gi],  a['l4'] if a else None, 1e-5)
        # long view: 2022 in slot 0, blank slot 1, then 2023-2026
        cmp(f'ov-long labels', e['llab'], ['2022','','2023','2024','2025','2026'])
        slots={2022:0,2023:2,2024:3,2025:4,2026:5}
        exp_p=[None]*6; exp_l=[None]*6
        for y,s in slots.items():
            a=A('city',['city'],bk,y,cat)
            if a: exp_p[s]=a['prof']; exp_l[s]=a['l1']
        for i in range(6):
            cmp(f'ov-long prof {e["st"]} slot{i}', e['long'][0]['v'][i], exp_p[i], 1e-5)
            cmp(f'ov-long l1 {e["st"]} slot{i}',   e['long'][1]['v'][i], exp_l[i], 1e-5)
    elif p=='di2':
        base,bk,mk=e['st']; key=MET[mk]
        # scatter: one dataset per borough, x=dL1 y=dProf
        pts={}
        for ds in e['scatter']:
            for pt in ds['v']: pts.setdefault(ds['l'],[]).append(pt)
        seen=0
        for b in BOROS:
            ds=[d for d in DISTRICTS if BORO_OF[d]==b]
            exp=[]
            for d in ds:
                cur=A('dist',[d],bk,2026,0); bse=A('dist',[d],bk,base,0)
                if cur and bse: exp.append((cur['l1']-bse['l1'], cur['prof']-bse['prof']))
            got=sorted([(pt['x'],pt['y']) for pt in pts.get(b,[])])
            exp=sorted(exp)
            cmp(f'di-scatter count {b} {e["st"]}', len(got), len(exp))
            for (gx,gy),(ex,ey) in zip(got,exp):
                cmp(f'di-scatter x {b} {e["st"]}', gx, ex, 1e-5)
                cmp(f'di-scatter y {b} {e["st"]}', gy, ey, 1e-5)
                seen+=1
        # ranked bars, sorted by change in the direction of improvement
        exp=[]
        for d in DISTRICTS:
            cur=A('dist',[d],bk,2026,0); bse=A('dist',[d],bk,base,0)
            if cur and bse: exp.append(('D'+d, cur[key]-bse[key]))
        exp.sort(key=lambda t:-t[1]*GOOD[mk])
        cmp(f'di-rank labels {e["st"]}', e['rlab'], [t[0] for t in exp])
        for i,(lab,v) in enumerate(exp):
            cmp(f'di-rank v {lab} {e["st"]}', e['rank'][0]['v'][i], v, 1e-5)
    elif p=='bo2':
        base,bk,mk=e['st']; key=MET[mk]
        # borough x grade table: two rows per borough (prof then L1)
        rows=e['tbl'][1:]
        gk=['g3','g4','g5','g6','g7','g8','all']
        for bi,b in enumerate(BOROS):
            rp=rows[bi*2]; rl=rows[bi*2+1]
            cmp(f'bo-tbl name {b}', rp[0], b)
            for j,g in enumerate(gk):
                a=A('boro',[b],g,2026,0)
                cmp(f'bo-tbl prof {b} {g}', rp[2+j], f1(a['prof']) if a else 's')
                cmp(f'bo-tbl l1 {b} {g}',   rl[1+j], f1(a['l1']) if a else 's')
        # district bars coloured by borough, ordered by the metric
        exp=[]
        for d in DISTRICTS:
            a=A('dist',[d],bk,2026,0)
            if a: exp.append(('D'+d, a[key]))
        exp.sort(key=lambda t:-t[1]*GOOD[mk])
        cmp(f'bo-dbar labels {e["st"]}', e['dlab'], [t[0] for t in exp])
        for i,(lab,v) in enumerate(exp):
            cmp(f'bo-dbar v {lab} {e["st"]}', e['dbar'][0]['v'][i], v, 1e-5)
    elif p=='ph2':
        mk,dim,cat=e['st']; cat=int(cat); key=MET[mk]
        specs=[('Elementary Phase 1',PHASE['elem1'],'35'),('Elementary Phase 2',PHASE['elem2'],'35'),
               ('Middle school Phase 1',PHASE['ms1'],'68'),
               ('Not yet in middle school rollout',[d for d in range(1,33) if d not in PHASE['ms1']],'68'),
               ('Elementary Phase 1 districts',PHASE['elem1'],'68'),
               ('Elementary Phase 2 districts',PHASE['elem2'],'68')]
        data=[r for r in e['tbl'] if len(r)>=9 and r[0] in [s[0] for s in specs]]
        cmp(f'ph-tbl rowcount {e["st"]}', len(data), 6)
        for r in data:
            spec=[s for s in specs if s[0]==r[0]]
            # two specs share a name prefix; match on district count too
            spec=[s for s in spec if str(len(s[1]))==r[1]] or spec
            name,ds,bk=spec[0]
            gids=['%02d'%d for d in ds]
            cmp(f'ph-tbl ndist {name} {e["st"]}', r[1], str(len(ds)))
            a26=A('dist',gids,bk,2026,cat)
            cmp(f'ph-tbl n {name} {e["st"]}', r[2], numf(a26['n']) if a26 else 's')
            for yi,y in enumerate(MODERN):
                a=A('dist',gids,bk,y,cat)
                cmp(f'ph-tbl {name} {y} {e["st"]}', r[3+yi], f1(a[key]) if a else 's')
            a23=A('dist',gids,bk,2023,cat); a24=A('dist',gids,bk,2024,cat)
            cmp(f'ph-tbl d23 {name} {e["st"]}', r[7], ppf(a26[key]-a23[key]) if a26 and a23 else 's')
            cmp(f'ph-tbl d24 {name} {e["st"]}', r[8], ppf(a26[key]-a24[key]) if a26 and a24 else 's')
    elif p=='sg2':
        PAIRS=[('White','Black'),('White','Hispanic'),('Not Econ Disadv','Econ Disadv'),
               ('Not SWD','SWD'),('Never ELL','Current ELL')]
        for i,(x,y2) in enumerate(PAIRS):
            for yi,y in enumerate(MODERN):
                Ax=A('city',['city'],'all',y,CATS.index(x)); Bx=A('city',['city'],'all',y,CATS.index(y2))
                cmp(f'sg-gap chart {x}-{y2} {y}', e['gapchart'][i]['v'][yi],
                    (Ax['prof']-Bx['prof']) if Ax and Bx else None, 1e-5)
    elif p=='tf2':
        mk=e['st'][0]; key=MET[mk]
        CBT={'3':None,'4':2025,'5':2024,'6':2025,'7':None,'8':2024}
        for gi,g in enumerate(['3','4','5','6','7','8']):
            r=e['tbl'][gi+1]
            cmp(f'tf-tbl grade {g}', r[0], f'Grade {g}')
            cmp(f'tf-tbl cbt {g}', r[1], str(CBT[g]) if CBT[g] else '—')
            vs=[]
            for yi,y in enumerate(MODERN):
                a=A('city',['city'],'g'+g,y,0); vs.append(a[key] if a else None)
                cmp(f'tf-tbl {g} {y} {mk}', r[2+yi], f1(a[key]) if a else 's')
            cmp(f'tf-tbl d {g} {mk}', r[6], ppf(vs[3]-vs[0]))
        GR={'5 and 8':['5','8'],'4 and 6':['4','6'],'3 and 7':['3','7']}
        for gi,(lab,gs) in enumerate(GR.items()):
            base=None
            for yi,y in enumerate(MODERN):
                n=c1=c2=c3=c4=0; wm=0.0
                for g in gs:
                    nt,ms,a,b,c,d=C['city'][('city',g,y,'All Students')]
                    n+=nt;c1+=a;c2+=b;c3+=c;c4+=d;wm+=ms*nt
                val={'prof':(c3+c4)/n*100,'l1':c1/n*100,'l4':c4/n*100,'mean':wm/n}[mk]
                if yi==0: base=val
                else: cmp(f'tf-delta {lab} {y} {mk}', e['delta'][gi]['v'][yi-1], val-base, 1e-5)

# ---------------- CSV exports ----------------
CSV=json.load(open(os.path.join(R,'audit/csv_dump.json')))
def cell(v): return '' if v is None else str(v)

# ov CSV was captured with group = Econ Disadv, all grades
rows=CSV['ov']; cat=CATS.index('Econ Disadv')
cmp('csv ov group', rows[2][1], 'Econ Disadv')
for i,y in enumerate(YEARS):
    r=rows[6+i]; a=A('city',['city'],'all',y,cat)
    cmp(f'csv ov year {y}', str(r[0]), str(y))
    cmp(f'csv ov n {y}', str(r[1]), str(a['n']))
    for j,k in enumerate(['l1','l2','l3','l4','prof']):
        cmp(f'csv ov {k} {y}', r[2+j], f2(a[k]))
    cmp(f'csv ov mean {y}', r[7], f1(a['mean']))
    cmp(f'csv ov comparable {y}', r[8], 'yes' if y>=2023 else 'no (previous standards)')

# di CSV: grades 6-8, all students, baseline 2024
rows=CSV['di']; bk='68'; base=2024
cmp('csv di grades', rows[1][1], BAND_LABEL[bk])
for i,d in enumerate(DISTRICTS):
    r=rows[7+i]; cur=A('dist',[d],bk,2026,0); bse=A('dist',[d],bk,base,0)
    cmp(f'csv di name {d}', r[0], f'District {d}')
    cmp(f'csv di boro {d}', r[1], BORO_OF[d])
    n=int(d)
    cmp(f'csv di elem {d}', r[2], 'Elem Phase 1' if n in PHASE['elem1'] else 'Elem Phase 2' if n in PHASE['elem2'] else '—')
    cmp(f'csv di ms {d}', r[3], 'MS Phase 1' if n in PHASE['ms1'] else 'MS Phase 2' if n in PHASE['ms2'] else '—')
    # cols: 0 District 1 Borough 2 ElemPhase 3 MSPhase 4 ReadsPL 5 ElemCurr
    #       6 MSCurr 7 SolvesPL 8 Tested 9-13 %L1..%L3+4 14 mean
    #       15 TestedBase 16 %L1Base 17 %ProfBase 18 dL1 19 dProf 20 Signal
    cmp(f'csv di n {d}', str(r[8]), str(cur['n']))
    for j,k in enumerate(['l1','l2','l3','l4','prof']):
        cmp(f'csv di {k} {d}', r[9+j], f2(cur[k]))
    cmp(f'csv di mean {d}', r[14], f1(cur['mean']))
    cmp(f'csv di nbase {d}', str(r[15]), str(bse['n']))
    cmp(f'csv di l1base {d}', r[16], f2(bse['l1']))
    cmp(f'csv di profbase {d}', r[17], f2(bse['prof']))
    cmp(f'csv di dl1 {d}', r[18], f'{cur["l1"]-bse["l1"]:.2f}')
    cmp(f'csv di dprof {d}', r[19], f'{cur["prof"]-bse["prof"]:.2f}')
    dl1=cur['l1']-bse['l1']; dpr=cur['prof']-bse['prof']
    cmp(f'csv di sig {d}', r[20],
        'Improved on both' if dl1<0 and dpr>0 else 'Mixed' if dl1<0 or dpr>0 else 'Worse on both')

# bo CSV: grades 3-5
rows=CSV['bo']; bk='35'
cmp('csv bo grades', rows[1][1], BAND_LABEL[bk])
cmp('csv bo header', rows[6][0], 'Borough')   # data starts at 7, after the header row
i=7
for b in BOROS:
    for y in MODERN:
        r=rows[i]; i+=1; a=A('boro',[b],bk,y,0)
        cmp(f'csv bo {b} {y} name', r[0], b); cmp(f'csv bo {b} {y} year', str(r[1]), str(y))
        cmp(f'csv bo {b} {y} n', str(r[2]), str(a['n']))
        for j,k in enumerate(['l1','l2','l3','l4','prof']):
            cmp(f'csv bo {b} {y} {k}', r[3+j], f2(a[k]))
        cmp(f'csv bo {b} {y} mean', r[8], f1(a['mean']))

# ph CSV: SWD
rows=CSV['ph']; cat=CATS.index('SWD')
cmp('csv ph group', rows[1][1], 'SWD')
specs=[('Elementary Phase 1',PHASE['elem1'],'35'),('Elementary Phase 2',PHASE['elem2'],'35'),
       ('Middle school Phase 1',PHASE['ms1'],'68'),
       ('Not yet in middle school rollout',[d for d in range(1,33) if d not in PHASE['ms1']],'68'),
       ('Elementary Phase 1 districts (grades 6-8 check)',PHASE['elem1'],'68'),
       ('Elementary Phase 2 districts (grades 6-8 check)',PHASE['elem2'],'68')]
i=6
for name,ds,bk in specs:
    gids=['%02d'%d for d in ds]
    for y in MODERN:
        r=rows[i]; i+=1; a=A('dist',gids,bk,y,cat)
        cmp(f'csv ph {name} {y} name', r[0], name)
        cmp(f'csv ph {name} {y} band', r[1], '3-5' if bk=='35' else '6-8')
        cmp(f'csv ph {name} {y} nd', str(r[2]), str(len(ds)))
        cmp(f'csv ph {name} {y} n', str(r[4]), str(a['n']))
        cmp(f'csv ph {name} {y} l1', r[5], f2(a['l1']))
        cmp(f'csv ph {name} {y} prof', r[6], f2(a['prof']))
        cmp(f'csv ph {name} {y} mean', r[7], f1(a['mean']))

# sg CSV: citywide, all grades, every category
rows=CSV['sg']; i=6
for c in range(len(CATS)):
    for y in MODERN:
        r=rows[i]; i+=1; a=A('city',['city'],'all',y,c)
        cmp(f'csv sg {CATS[c]} {y} name', r[0], CATS[c])
        cmp(f'csv sg {CATS[c]} {y} n', str(r[2]), str(a['n']) if a else 's')
        for j,k in enumerate(['l1','l2','l3','l4','prof']):
            cmp(f'csv sg {CATS[c]} {y} {k}', r[3+j], f2(a[k]) if a else 's')
        cmp(f'csv sg {CATS[c]} {y} mean', r[8], f1(a['mean']) if a else 's')

# tf CSV: group block then per-grade block
rows=CSV['tf']; i=6
GR=[('Moved to computer in 2024',['5','8']),('Moved to computer in 2025',['4','6']),
    ('Not named in the transition',['3','7'])]
for name,gs in GR:
    for y in MODERN:
        r=rows[i]; i+=1
        n=c1=c3=c4=0; wm=0.0
        for g in gs:
            nt,ms,a1,b1,c,d=C['city'][('city',g,y,'All Students')]
            n+=nt;c1+=a1;c3+=c;c4+=d;wm+=ms*nt
        cmp(f'csv tf {name} {y} name', r[0], name)
        cmp(f'csv tf {name} {y} n', str(r[3]), str(n))
        cmp(f'csv tf {name} {y} l1', r[4], f2(c1/n*100))
        cmp(f'csv tf {name} {y} prof', r[5], f2((c3+c4)/n*100))
        cmp(f'csv tf {name} {y} mean', r[6], f1(wm/n))
i+=2
CBT={'3':'not stated','4':'2025','5':'2024','6':'2025','7':'not stated','8':'2024'}
for g in ['3','4','5','6','7','8']:
    for y in MODERN:
        r=rows[i]; i+=1; a=A('city',['city'],'g'+g,y,0)
        cmp(f'csv tf grade {g} {y} name', r[0], f'Grade {g}')
        cmp(f'csv tf grade {g} {y} cbt', str(r[1]), CBT[g])
        cmp(f'csv tf grade {g} {y} n', str(r[3]), str(a['n']))
        cmp(f'csv tf grade {g} {y} l1', r[4], f2(a['l1']))
        cmp(f'csv tf grade {g} {y} prof', r[5], f2(a['prof']))
        cmp(f'csv tf grade {g} {y} mean', r[6], f1(a['mean']))

print('='*72)
print('PASS 2 — charts, tables and CSV exports not covered by the first render audit')
print('='*72)
print(f'values checked : {checked:,}')
print(f'MISMATCHES     : {len(bad):,}')
if bad:
    for t,g,x in bad[:30]: print('  ',t,'| shown:',repr(g),'| expected:',repr(x))
    sys.exit(1)
print('\nRESULT: PASS')
