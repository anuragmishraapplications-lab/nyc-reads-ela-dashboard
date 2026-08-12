"""Extract NYC ELA results into a compact JSON payload for the dashboard.

Design notes (audited):
  * The raw files are internally exact: #L1+#L2+#L3+#L4 == Number Tested,
    %Lx == #Lx / Number Tested * 100 (to float precision), #L3+4 == #L3+#L4.
    Therefore we store COUNTS ONLY and derive every percentage in the browser.
    This guarantees that aggregation (across districts / across grades) is exact
    and that no displayed percentage is a rounded-input-of-a-rounded-input.
  * Suppressed cells ('s') are dropped; the UI renders them as 's'.
"""
import openpyxl, json, os, sys

YEARS  = [2018, 2019, 2022, 2023, 2024, 2025, 2026]
GRADES = ['3', '4', '5', '6', '7', '8', 'All Grades']
YI = {y: i for i, y in enumerate(YEARS)}
GI = {g: i for i, g in enumerate(GRADES)}

SHEETS = ['ELA - All', 'ELA - SWD', 'ELA - Ethnicity', 'ELA - Gender',
          'ELA - Econ Status', 'ELA - ELL', 'ELA - Gender by Ethnicity']

# canonical category ordering -> index
CATS = ['All Students',
        'SWD', 'Not SWD',
        'Econ Disadv', 'Not Econ Disadv',
        'Current ELL', 'Ever ELL', 'Never ELL',
        'Asian', 'Black', 'Hispanic', 'Multi-Racial', 'Native American', 'White',
        'Female', 'Male', 'Neither Female nor Male']
CI = {c: i for i, c in enumerate(CATS)}

# which sheet a category belongs to (dimension grouping for the UI)
DIMS = [
    ('all',   'All students',          ['All Students']),
    ('swd',   'Students with disabilities', ['SWD', 'Not SWD']),
    ('econ',  'Economic status',       ['Econ Disadv', 'Not Econ Disadv']),
    ('ell',   'English language learners', ['Current ELL', 'Ever ELL', 'Never ELL']),
    ('eth',   'Race / ethnicity',      ['Asian', 'Black', 'Hispanic', 'Multi-Racial',
                                        'Native American', 'White']),
    ('gen',   'Gender',                ['Female', 'Male', 'Neither Female nor Male']),
]

BORO_OF = {}
for d in range(1, 7):   BORO_OF['%02d' % d] = 'Manhattan'
for d in range(7, 13):  BORO_OF['%02d' % d] = 'Bronx'
for d in list(range(13, 24)) + [32]: BORO_OF['%02d' % d] = 'Brooklyn'
for d in range(24, 31): BORO_OF['%02d' % d] = 'Queens'
BORO_OF['31'] = 'Staten Island'

BOROS = ['Bronx', 'Brooklyn', 'Manhattan', 'Queens', 'Staten Island']
BORO_KEY = {'BRONX': 'Bronx', 'BROOKLYN': 'Brooklyn', 'MANHATTAN': 'Manhattan',
            'QUEENS': 'Queens', 'STATEN ISLAND': 'Staten Island'}
DISTRICTS = ['%02d' % d for d in range(1, 33)]

# NYC Reads launch timeline (source: NYCPS "NYC Reads phase info", verified
# against the screenshot supplied by the team on 2026-08-11).
PHASE = {
    'elem1': [5, 11, 12, 14, 16, 19, 20, 21, 22, 23, 25, 26, 29, 30, 32],   # +75 (not in files)
    'elem2': [1, 2, 3, 4, 6, 7, 8, 9, 10, 13, 15, 17, 18, 24, 27, 28, 31],  # +75
    'ms1':   [1, 3, 7, 9, 11, 12, 13, 19],
    'ms2':   [2, 4, 8, 10, 14, 18, 20, 26, 29, 32],                          # +75
}

# ---------------------------------------------------------------------------
# NYC Reads / NYC Solves vendor and curriculum assignments, by district.
# Source: "2026-27 District Reads Solves Data (Draft).xlsx", sheet
# "District Sustainability Conditi".
#
# ONLY the vendor, curriculum and year-joined columns are read. The source
# sheet also carries superintendent and deputy names, their email addresses,
# and internal district-condition narrative. None of that is loaded, and the
# source workbook is kept out of version control (see .gitignore), because the
# published dashboard is public.
# ---------------------------------------------------------------------------
VENDOR_SRC = 'private/2026-27 District Reads Solves Data (Draft).xlsx'
VENDOR_SHEET = 'District Sustainability Conditi'
VENDOR_COLS = ('Reads Vendor', 'Solves Vendor', 'NYC Reads Elementary Curriculum',
               'NYC Reads MS Curriculum', 'NYC Solves Curriculum')
BANNED_COLS = ('Superintendent', 'Supt Email', 'Dept Supt', 'Dep Supt Email', 'Supt')

stats = {'read': 0, 'kept': 0, 'suppressed': 0, 'bad': 0}
CHECKS = []          # populated by read_level / main; every one must pass
INVENTORY = []


DEV = {'pct': 0.0, 'p34': 0.0, 'sum100': 0.0}


def read_level(path, idcol):
    """Return list of (idvalue, gradeIdx, yearIdx, catIdx, nTested, mean, c1,c2,c3,c4)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    local = {'loaded': 0, 'suppressed': 0}
    for sn in wb.sheetnames:
        if sn not in SHEETS:
            continue
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        hdr = list(next(rows))
        H = {h: i for i, h in enumerate(hdr)}
        need = ['Grade', 'Year', 'Category', 'Number Tested', 'Mean Scale Score',
                '# Level 1', '# Level 2', '# Level 3', '# Level 4', '# Level 3+4',
                '% Level 1', '% Level 2', '% Level 3', '% Level 4', '% Level 3+4']
        missing = [c for c in need if c not in H]
        assert not missing, (path, sn, missing)
        for r in rows:
            if r is None or r[0] is None:
                continue
            stats['read'] += 1
            cat = r[H['Category']]
            if cat not in CI:                       # e.g. Gender-by-Ethnicity combos
                continue
            vals = [r[H[c]] for c in ['Number Tested', 'Mean Scale Score', '# Level 1',
                                      '# Level 2', '# Level 3', '# Level 4']]
            if any(v == 's' for v in vals):
                stats['suppressed'] += 1
                local['suppressed'] += 1
                continue
            if any(not isinstance(v, (int, float)) for v in vals):
                stats['bad'] += 1
                continue
            nt, ms, c1, c2, c3, c4 = vals
            # hard invariants -- refuse to emit a row that violates them
            assert c1 + c2 + c3 + c4 == nt, (path, sn, r)
            assert r[H['# Level 3+4']] == c3 + c4, (path, sn, r)
            if nt:
                for c, pk in zip((c1, c2, c3, c4),
                                 ('% Level 1', '% Level 2', '% Level 3', '% Level 4')):
                    d = abs(c / nt * 100 - r[H[pk]])
                    DEV['pct'] = max(DEV['pct'], d)
                    assert d < 1e-4, (path, sn, pk, r)
                d = abs((c3 + c4) / nt * 100 - r[H['% Level 3+4']])
                DEV['p34'] = max(DEV['p34'], d)
                assert d < 1e-4, (path, sn, r)
                d = abs(sum(r[H[k]] for k in ('% Level 1', '% Level 2',
                                              '% Level 3', '% Level 4')) - 100)
                DEV['sum100'] = max(DEV['sum100'], d)
                assert d < 1e-3, (path, sn, r)
            gi = GI[str(r[H['Grade']])]
            yi = YI[r[H['Year']]]
            idv = r[H[idcol]] if idcol else None
            out.append([idv, gi, yi, CI[cat], int(nt), round(float(ms), 6),
                        int(c1), int(c2), int(c3), int(c4)])
            stats['kept'] += 1
            local['loaded'] += 1
    INVENTORY.append((os.path.basename(path), local['loaded'], local['suppressed']))
    return out


def pack(rows, geo_index):
    """[geoIdx, gradeIdx, yearIdx, catIdx, n, mean, c1,c2,c3,c4]"""
    packed = []
    for idv, gi, yi, ci, nt, ms, c1, c2, c3, c4 in rows:
        g = 0 if geo_index is None else geo_index[idv]
        packed.append([g, gi, yi, ci, nt, ms, c1, c2, c3, c4])
    packed.sort()
    return packed


def read_vendors(base):
    """District -> provider / curriculum for SY 2025-26.

    Source: the curriculum-and-JESP chart supplied by the team on 11 Aug 2026,
    transcribed into data/vendors_sy2025_26.json. That is the school year the
    spring 2026 test measures, so it is the correct vintage for these results;
    the 2026-27 draft workbook describes assignments that begin AFTER the 2026
    test and is no longer used for provider or curriculum.

    JESP is the job-embedded support provider, i.e. the professional learning
    vendor. K-5 and grades 6-8 are held separately because they differ.

    The chart also carries a Phase column. It disagrees with the NYCPS launch
    timeline for seven districts, and the timeline is corroborated by the
    "year joined" columns of the 2026-27 workbook for all 32, so the phase
    assignments are NOT taken from the chart. The disagreement is carried into
    the payload so the dashboard can show it rather than hide it.
    """
    path = os.path.join(base, 'vendors_sy2025_26.json')
    if not os.path.exists(path):
        print('WARNING: vendor source not found, provider page will be empty')
        return None
    with open(path) as f:
        src = json.load(f)
    recs = src['districts']
    assert len(recs) == 32, 'expected 32 districts, got %d' % len(recs)

    def roster(key):
        return sorted({recs[str(d)][key] for d in range(1, 33) if recs[str(d)][key]})
    kj, kc = roster('k5Jesp'), roster('k5Curr')
    mj, mc = roster('msJesp'), roster('msCurr')

    def idx(r, v):
        return r.index(v) if v else None
    by = [{'kj': idx(kj, recs[str(d)]['k5Jesp']), 'kc': idx(kc, recs[str(d)]['k5Curr']),
           'mj': idx(mj, recs[str(d)]['msJesp']), 'mc': idx(mc, recs[str(d)]['msCurr'])}
          for d in range(1, 33)]

    official = {**{d: 1 for d in PHASE['elem1']}, **{d: 2 for d in PHASE['elem2']}}
    chart = {int(k): v for k, v in src['_phaseInChart'].items()}
    disagree = [{'d': d, 'timeline': official[d], 'chart': chart[d]}
                for d in range(1, 33) if official[d] != chart[d]]

    payload = {'source': 'SY 2025-26',
               'k5JespRoster': kj, 'k5CurrRoster': kc,
               'msJespRoster': mj, 'msCurrRoster': mc,
               'byDistrict': by, 'phaseDisagree': disagree}
    blob = json.dumps(payload)
    assert '@' not in blob, 'an email address reached the vendor payload'
    print('vendors (SY2025-26): %d K-5 providers, %d K-5 curricula, %d grades 6-8 providers; '
          'phase column disagrees with the launch timeline for %d districts'
          % (len(kj), len(kc), len(mj), len(disagree)))
    return payload


def main():
    D = os.path.join(os.path.dirname(__file__), '..', 'data')
    city = read_level(os.path.join(D, 'citywide-ela-results-public.xlsx'), None)
    boro = read_level(os.path.join(D, 'borough-ela-results-public.xlsx'), 'Borough')
    dist = read_level(os.path.join(D, 'district-ela-results-public.xlsx'), 'District')

    boro = [[BORO_KEY[r[0]]] + r[1:] for r in boro]
    bidx = {b: i for i, b in enumerate(BOROS)}
    didx = {d: i for i, d in enumerate(DISTRICTS)}

    packed = {'city': pack(city, None), 'boro': pack(boro, bidx), 'dist': pack(dist, didx)}

    # ---- check: 'All Grades' row equals the sum of grades 3-8, wherever the
    #      six grade rows are all present (i.e. none suppressed).
    ag_bad = ag_tot = ag_part = 0
    for lvl, rows in packed.items():
        groups = {}
        for r in rows:
            groups.setdefault((r[0], r[2], r[3]), {})[r[1]] = r
        for k, g in groups.items():
            if GI['All Grades'] not in g:
                continue
            ag_tot += 1
            parts = [g[x] for x in range(6) if x in g]
            if len(parts) < 6:
                ag_part += 1
                continue
            for col in (4, 6, 7, 8, 9):
                if sum(p[col] for p in parts) != g[GI['All Grades']][col]:
                    ag_bad += 1
                    break
    assert ag_bad == 0, 'All Grades != sum(3-8) in %d groups' % ag_bad

    # ---- check: every district x year x grade cell exists for All Students
    exp = {(d, y, g) for d in range(32) for y in range(len(YEARS)) for g in range(7)}
    have = {(r[0], r[2], r[1]) for r in packed['dist'] if r[3] == CI['All Students']}
    missing = len(exp - have)
    assert missing == 0, 'missing district cells: %d' % missing

    # ---- reconciliation between the three files (All Students, All Grades)
    def total(rows, year):
        yi = YI[year]
        return sum(r[4] for r in rows
                   if r[1] == GI['All Grades'] and r[2] == yi and r[3] == CI['All Students'])
    recon = {str(y): {'city': total(packed['city'], y),
                      'boro': total(packed['boro'], y),
                      'dist': total(packed['dist'], y)} for y in YEARS}

    checks = [
        {'name': 'Level counts sum to the number of students tested',
         'scope': 'every loaded row',
         'pass': True,
         'detail': '# Level 1 + 2 + 3 + 4 = Number Tested, exactly, in all %s rows.' % f"{stats['kept']:,}"},
        {'name': 'Published percentages match percentages recomputed from counts',
         'scope': 'every loaded row',
         'pass': True,
         'detail': 'Largest disagreement across all four levels: %.2e percentage points.' % DEV['pct']},
        {'name': 'Published % Level 3+4 matches Level 3 plus Level 4',
         'scope': 'every loaded row',
         'pass': True,
         'detail': 'Largest disagreement: %.2e percentage points; the counts agree exactly.' % DEV['p34']},
        {'name': 'The four level percentages sum to 100',
         'scope': 'every loaded row',
         'pass': True,
         'detail': 'Largest departure from 100: %.2e percentage points.' % DEV['sum100']},
        {'name': 'The All Grades row equals the sum of grades 3 to 8',
         'scope': 'every geography, year and student group',
         'pass': ag_bad == 0,
         'detail': '%s groups checked, %s failures. %s groups have at least one suppressed grade row; '
                   'for those the tool reads the All Grades row rather than summing.'
                   % (f'{ag_tot:,}', ag_bad, f'{ag_part:,}')},
        {'name': 'No district, year or grade is missing',
         'scope': 'district file, all students',
         'pass': missing == 0,
         'detail': '32 districts x 7 years x 7 grade rows = 1,568 cells, all present.'},
        {'name': 'The three files do not reconcile with each other',
         'scope': 'citywide vs borough vs district totals',
         'pass': True,
         'detail': 'Expected, and handled: in 2026 the district file totals %s tested students, the borough '
                   'file %s and the citywide file %s. Each level is read only from its own file.'
                   % (f"{recon['2026']['dist']:,}", f"{recon['2026']['boro']:,}", f"{recon['2026']['city']:,}")},
    ]

    inv_meta = {'citywide-ela-results-public.xlsx': ('2018–2026', 'New York City'),
                'borough-ela-results-public.xlsx': ('2018–2026', '5 boroughs'),
                'district-ela-results-public.xlsx': ('2018–2026', '32 community school districts')}
    inventory = [{'file': f, 'loaded': ld, 'suppressed': sp,
                  'years': inv_meta[f][0], 'geos': inv_meta[f][1],
                  'cats': len(CATS)} for f, ld, sp in INVENTORY]

    payload = {
        'years': YEARS,
        'grades': GRADES,
        'cats': CATS,
        'dims': [{'k': k, 'label': lab, 'cats': [CI[c] for c in cs]} for k, lab, cs in DIMS],
        'boros': BOROS,
        'districts': DISTRICTS,
        'districtBoro': [BORO_OF[d] for d in DISTRICTS],
        'phase': PHASE,
        'vendors': read_vendors(D),
        'meta': {
            'rowsRead': stats['read'], 'rowsChecked': stats['kept'],
            'suppressed': stats['suppressed'],
            'checks': checks, 'inventory': inventory, 'recon': recon,
        },
        'city': packed['city'], 'boro': packed['boro'], 'dist': packed['dist'],
    }
    outp = os.path.join(D, 'payload.json')
    with open(outp, 'w') as f:
        json.dump(payload, f, separators=(',', ':'))
    print('stats', stats)
    print('max deviations', {k: '%.2e' % v for k, v in DEV.items()})
    print('AllGrades check: %d groups, %d failures, %d partial' % (ag_tot, ag_bad, ag_part))
    print('recon 2026', recon['2026'])
    print('rows city/boro/dist:', len(payload['city']), len(payload['boro']), len(payload['dist']))
    print('bytes', os.path.getsize(outp))
    print('ALL CHECKS PASSED')


if __name__ == '__main__':
    main()
