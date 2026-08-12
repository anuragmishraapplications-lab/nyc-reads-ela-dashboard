"""Audit what the dashboard RENDERS, against an independent recomputation.

Reads the original .xlsx files (again, from scratch) and re-derives every KPI,
chart point and table cell captured from the live page, including the direction
and sign of every change.  Any disagreement is reported.
"""
import openpyxl, json, os, sys, re
R = os.path.join(os.path.dirname(__file__), '..')
YEARS=[2018,2019,2022,2023,2024,2025,2026]; MODERN=[2023,2024,2025,2026]
CATS=['All Students','SWD','Not SWD','Econ Disadv','Not Econ Disadv','Current ELL','Ever ELL',
      'Never ELL','Asian','Black','Hispanic','Multi-Racial','Native American','White',
      'Female','Male','Neither Female nor Male']
BOROS=['Bronx','Brooklyn','Manhattan','Queens','Staten Island']
BKEY={'BRONX':'Bronx','BROOKLYN':'Brooklyn','MANHATTAN':'Manhattan','QUEENS':'Queens','STATEN ISLAND':'Staten Island'}
DISTRICTS=['%02d'%d for d in range(1,33)]
BANDS={'all':['All Grades'],'35':['3','4','5'],'68':['6','7','8'],
       'g3':['3'],'g4':['4'],'g5':['5'],'g6':['6'],'g7':['7'],'g8':['8']}
BORO_OF={}
for d in range(1,7): BORO_OF['%02d'%d]='Manhattan'
for d in range(7,13): BORO_OF['%02d'%d]='Bronx'
for d in list(range(13,24))+[32]: BORO_OF['%02d'%d]='Brooklyn'
for d in range(24,31): BORO_OF['%02d'%d]='Queens'
BORO_OF['31']='Staten Island'
PHASE={'elem1':[5,11,12,14,16,19,20,21,22,23,25,26,29,30,32],
       'elem2':[1,2,3,4,6,7,8,9,10,13,15,17,18,24,27,28,31],
       'ms1':[1,3,7,9,11,12,13,19],'ms2':[2,4,8,10,14,18,20,26,29,32]}

def load(path, idcol):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True); C={}
    for sn in wb.sheetnames:
        if not sn.startswith('ELA'): continue
        ws=wb[sn]; it=ws.iter_rows(values_only=True); hdr=list(next(it)); H={h:i for i,h in enumerate(hdr)}
        for r in it:
            if r is None or r[0] is None: continue
            cat=r[H['Category']]
            if cat not in CATS: continue
            v=[r[H[c]] for c in ('Number Tested','Mean Scale Score','# Level 1','# Level 2','# Level 3','# Level 4')]
            if any(x=='s' for x in v): continue
            gid = BKEY[r[H[idcol]]] if idcol=='Borough' else (r[H[idcol]] if idcol else 'city')
            C[(gid,str(r[H['Grade']]),r[H['Year']],cat)] = v
    return C
C={'city':load(os.path.join(R,'data','citywide-ela-results-public.xlsx'),None),
   'boro':load(os.path.join(R,'data','borough-ela-results-public.xlsx'),'Borough'),
   'dist':load(os.path.join(R,'data','district-ela-results-public.xlsx'),'District')}

def A(lvl, gids, bk, year, cat):
    n=c1=c2=c3=c4=0; wm=0.0; found=False
    for gid in gids:
        for g in BANDS[bk]:
            v=C[lvl].get((gid,g,year,CATS[cat] if isinstance(cat,int) else cat))
            if v is None: continue
            found=True; nt,ms,a,b,c,d=v
            n+=nt;c1+=a;c2+=b;c3+=c;c4+=d;wm+=ms*nt
    if not found or n==0: return None
    return dict(n=n,l1=c1/n*100,l2=c2/n*100,l3=c3/n*100,l4=c4/n*100,prof=(c3+c4)/n*100,mean=wm/n,l4v=c4/n*100)
MET={'prof':'prof','l1':'l1','l4':'l4','mean':'mean'}

from decimal import Decimal, ROUND_HALF_UP
def _r1(v):
    # JS toFixed(1) rounds ties away from zero; Python's format rounds ties to
    # even.  Match the browser so a value like 31.25 is compared as 31.3.
    return str(Decimal(repr(float(v))).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
def f1(v): return '—' if v is None else _r1(v)
def ppf(v):
    if v is None: return '—'
    if abs(v) < 0.05: return '0.0'        # never render a signed zero
    return ('+' if v>0 else '−')+_r1(abs(v))
def numf(v): return f'{v:,}'

REN=json.load(open(os.path.join(R,'audit/render_dump.json')))
bad=[]; checked=0
def cmp(tag, got, exp, tol=None):
    global checked; checked+=1
    if tol is not None:
        if got is None and exp is None: return
        if got is None or exp is None: bad.append((tag,got,exp)); return
        if abs(got-exp)>tol: bad.append((tag,got,exp))
    else:
        if str(got)!=str(exp): bad.append((tag,got,exp))

for e in REN:
    p=e['p']
    if p=='ov':
        base,bk,dim,cat=e['st']; cat=int(cat)
        cur=A('city',['city'],bk,2026,cat); bse=A('city',['city'],bk,base,cat)
        k=e['kpi']
        cmp(f'ov kpi prof {e["st"]}', k[0], (f1(cur['prof']) if cur else '—')+'%')
        cmp(f'ov kpi dprof {e["st"]}', k[1], (ppf(cur['prof']-bse['prof']) if cur and bse else '—')+'pp')
        cmp(f'ov kpi l1 {e["st"]}', k[2], (f1(cur['l1']) if cur else '—')+'%')
        cmp(f'ov kpi dl1 {e["st"]}', k[3], (ppf(cur['l1']-bse['l1']) if cur and bse else '—')+'pp')
        cmp(f'ov kpi n {e["st"]}', k[4], numf(cur['n']) if cur else '—')
        gg=sum(1 for d in DISTRICTS
               if (lambda a,b: a and b and a['l1']<b['l1'] and a['prof']>b['prof'])(A('dist',[d],bk,2026,cat),A('dist',[d],bk,base,cat)))
        tot=sum(1 for d in DISTRICTS if A('dist',[d],bk,2026,cat) and A('dist',[d],bk,base,cat))
        cmp(f'ov kpi gg {e["st"]}', k[5], f'{gg}/{tot}')
        for i,y in enumerate(MODERN):
            a=A('city',['city'],bk,y,cat)
            cmp(f'ov trend prof {e["st"]} {y}', e['trend'][0]['v'][i], a['prof'] if a else None, 1e-5)
            cmp(f'ov trend l1 {e["st"]} {y}',   e['trend'][1]['v'][i], a['l1'] if a else None, 1e-5)
            for li,key in enumerate(['l1','l2','l3','l4']):
                cmp(f'ov dist {e["st"]} {y} {key}', e['dist'][li]['v'][i], a[key] if a else None, 1e-5)
        for gi,g in enumerate(['g3','g4','g5','g6','g7','g8']):
            a=A('city',['city'],g,2026,cat); b=A('city',['city'],g,base,cat)
            cmp(f'ov grades dprof {e["st"]} {g}', e['grades'][0]['v'][gi], (a['prof']-b['prof']) if a and b else None, 1e-5)
            cmp(f'ov grades dl1 {e["st"]} {g}',   e['grades'][1]['v'][gi], (a['l1']-b['l1']) if a and b else None, 1e-5)
    elif p=='di':
        base,bk,dim,cat=e['st']; cat=int(cat)
        for row in e['rows']:
            d=row[0].replace('District ','')
            cur=A('dist',[d],bk,2026,cat); bse=A('dist',[d],bk,base,cat)
            cmp(f'di boro {d}', row[1], BORO_OF[d])
            n=int(d)
            cmp(f'di elem {d}', row[2], 'Elem Phase 1' if n in PHASE['elem1'] else 'Elem Phase 2' if n in PHASE['elem2'] else '—')
            cmp(f'di ms {d}', row[3], 'MS Phase 1' if n in PHASE['ms1'] else 'MS Phase 2' if n in PHASE['ms2'] else '—')
            # 0 name 1 boro 2 elem 3 ms 4 k5prov 5 k5curr 6 tested 7 dTested%
            # 8 %L3-4 9 dL3-4 10 %L1 11 dL1 12 %L4 13 mean 14 signal
            cmp(f'di n {d} {e["st"]}', row[6], numf(cur['n']) if cur else 's')
            dn = (cur['n']-bse['n'])/bse['n']*100 if cur and bse and bse['n'] else None
            cmp(f'di dn {d} {e["st"]}', row[7], ppf(dn) if dn is not None else 's')
            cmp(f'di prof {d} {e["st"]}', row[8], f1(cur['prof']) if cur else 's')
            cmp(f'di dprof {d} {e["st"]}', row[9], ppf(cur['prof']-bse['prof']) if cur and bse else 's')
            cmp(f'di l1 {d} {e["st"]}', row[10], f1(cur['l1']) if cur else 's')
            cmp(f'di dl1 {d} {e["st"]}', row[11], ppf(cur['l1']-bse['l1']) if cur and bse else 's')
            cmp(f'di l4 {d} {e["st"]}', row[12], f1(cur['l4']) if cur else 's')
            cmp(f'di mean {d} {e["st"]}', row[13], f1(cur['mean']) if cur else 's')
            if cur and bse:
                dl1=cur['l1']-bse['l1']; dpr=cur['prof']-bse['prof']
                sig='Improved on both' if dl1<0 and dpr>0 else 'Mixed' if dl1<0 or dpr>0 else 'Worse on both'
            else: sig='—'
            cmp(f'di sig {d} {e["st"]}', row[14], sig)
    elif p=='bo':
        base,bk,m=e['st']; key=MET[m]
        for bi,b in enumerate(BOROS):
            for yi,y in enumerate(MODERN):
                a=A('boro',[b],bk,y,0)
                cmp(f'bo trend {b} {y} {e["st"]}', e['trend'][bi]['v'][yi], a[key] if a else None, 1e-5)
            a26=A('boro',[b],bk,2026,0)
            for li,k2 in enumerate(['l1','l2','l3','l4']):
                cmp(f'bo dist {b} {k2} {e["st"]}', e['dist'][li]['v'][bi], a26[k2] if a26 else None, 1e-5)
        got={lab:v for lab,v in zip([l for l in e['dbar'][0]['v']],[0])} if False else None
    elif p=='ph':
        m,dim,cat=e['st']; cat=int(cat); key=MET[m]
        specs=[('elem',[('elem1','35'),('elem2','35')]),
               ('ms',[('ms1','68'),('__not_ms1','68')]),
               ('pl',[('elem1','68'),('elem2','68')])]
        for field,gs in specs:
            for gi,(pk,bk) in enumerate(gs):
                ds = [d for d in range(1,33) if d not in PHASE['ms1']] if pk=='__not_ms1' else PHASE[pk]
                gids=['%02d'%d for d in ds]
                for yi,y in enumerate(MODERN):
                    a=A('dist',gids,bk,y,cat)
                    cmp(f'ph {field} {pk} {y} {e["st"]}', e[field][gi]['v'][yi], a[key] if a else None, 1e-5)
    elif p=='sg':
        lvl,dim,m=e['st']; key=MET[m]
        gid={'city':'city','boro':BOROS[1],'dist':DISTRICTS[6]}[lvl]
        DIMC={'eth':['Asian','Black','Hispanic','Multi-Racial','Native American','White'],
              'ell':['Current ELL','Ever ELL','Never ELL'],'swd':['SWD','Not SWD'],
              'gen':['Female','Male','Neither Female nor Male']}[dim]
        for ci,cn in enumerate(DIMC):
            for yi,y in enumerate(MODERN):
                a=A(lvl,[gid],'all',y,CATS.index(cn))
                cmp(f'sg trend {lvl} {cn} {y} {m}', e['trend'][ci]['v'][yi], a[key] if a else None, 1e-5)
        for row in e['tbl'][1:]:
            cn=row[0]; a=A(lvl,[gid],'all',2026,CATS.index(cn))
            cmp(f'sg tbl n {lvl} {cn}', row[1], numf(a['n']) if a else 's')
            cmp(f'sg tbl l1 {lvl} {cn}', row[2], f1(a['l1']) if a else 's')
            cmp(f'sg tbl prof {lvl} {cn}', row[3], f1(a['prof']) if a else 's')
        PAIRS=[('White','Black'),('White','Hispanic'),('Not Econ Disadv','Econ Disadv'),
               ('Not SWD','SWD'),('Never ELL','Current ELL')]
        for ri,(x,y2) in enumerate(PAIRS):
            row=e['gap'][ri+1]
            for yi,y in enumerate(MODERN):
                Ax=A('city',['city'],'all',y,CATS.index(x)); Bx=A('city',['city'],'all',y,CATS.index(y2))
                cmp(f'sg gap {x}-{y2} {y}', row[yi+1], f1(Ax['prof']-Bx['prof']) if Ax and Bx else 's')
    elif p=='tf':
        m=e['st'][0]; key=MET[m]
        GR={'y24':['5','8'],'y25':['4','6'],'pap':['3','7']}
        for gi,gk in enumerate(['y24','y25','pap']):
            for yi,y in enumerate(MODERN):
                n=c1=c2=c3=c4=0; wm=0.0
                for g in GR[gk]:
                    v=C['city'].get(('city',g,y,'All Students')); nt,ms,a,b,c,d=v
                    n+=nt;c1+=a;c2+=b;c3+=c;c4+=d;wm+=ms*nt
                val={'prof':(c3+c4)/n*100,'l1':c1/n*100,'l4':c4/n*100,'mean':wm/n}[m]
                cmp(f'tf trend {gk} {y} {m}', e['trend'][gi]['v'][yi], val, 1e-5)

    elif p=='ve':
        fk,bk,mk,minsize=e['st']; key=MET[mk]; base=2024
        VEND=json.load(open(os.path.join(R,'data/payload.json')))['vendors']
        FIELD={'k5j':('k5JespRoster','kj'), 'k5c':('k5CurrRoster','kc'),
               'msj':('msJespRoster','mj'), 'msc':('msCurrRoster','mc')}
        rosterKey, fieldKey = FIELD[fk]
        roster=VEND[rosterKey]
        groups=[(v,[i for i in range(32) if VEND['byDistrict'][i][fieldKey] is not None
                    and roster[VEND['byDistrict'][i][fieldKey]]==v]) for v in roster]
        groups=[(n,ds) for n,ds in groups if ds]
        good={'prof':1,'l1':-1,'l4':1,'mean':1}[mk]
        charted=[]
        for n,ds in groups:
            if len(ds) < int(minsize): continue
            gids=['%02d'%(i+1) for i in ds]
            cur=A('dist',gids,bk,2026,0); bse=A('dist',gids,bk,base,0)
            if cur and bse: charted.append((n,ds,gids,cur,bse))
        charted.sort(key=lambda r:-r[3][key]*good)
        cmp(f've-slope labels {e["st"]}', e['slab'], [f'{n} ({len(ds)})' for n,ds,_,_,_ in charted])
        for i,(n,ds,gids,cur,bse) in enumerate(charted):
            pair=e['slope'][0]['v'][i]
            cmp(f've-slope from {n} {e["st"]}', pair[0], bse[key], 1e-5)
            cmp(f've-slope to {n} {e["st"]}',   pair[1], cur[key], 1e-5)
        allrows=[]
        for n,ds in groups:
            gids=['%02d'%(i+1) for i in ds]
            cur=A('dist',gids,bk,2026,0); bse=A('dist',gids,bk,base,0)
            if cur: allrows.append((n,ds,gids,cur,bse))
        allrows.sort(key=lambda r:-r[3][key]*good)
        for ri,(n,ds,gids,cur,bse) in enumerate(allrows):
            trow=e['tbl'][ri+1]
            cmp(f've-tbl name {n} {e["st"]}', trow[0].replace(' (not charted)',''), n)
            cmp(f've-tbl nd {n} {e["st"]}', trow[1], str(len(ds)))
            cmp(f've-tbl n {n} {e["st"]}', trow[2], numf(cur['n']))
            for yi,y in enumerate(MODERN):
                a=A('dist',gids,bk,y,0)
                cmp(f've-tbl {n} {y} {e["st"]}', trow[3+yi], f1(a[key]) if a else 's')
            cmp(f've-tbl d {n} {e["st"]}', trow[7], ppf(cur[key]-bse[key]) if bse else 's')

print('='*72); print('RENDER AUDIT — values shown on the page vs independent recomputation'); print('='*72)
print(f'rendered values checked : {checked:,}')
print(f'MISMATCHES              : {len(bad):,}')
if bad:
    for t,g,x in bad[:30]: print('  ',t,'| shown:',g,'| expected:',x)
    sys.exit(1)
print('\nRESULT: PASS')
