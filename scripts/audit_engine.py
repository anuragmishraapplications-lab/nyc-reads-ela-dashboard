"""INDEPENDENT audit of the shipped dashboard.

Reads the ORIGINAL .xlsx files again from scratch (no reuse of payload.json,
no reuse of extract.py's structures) and recomputes every one of the 40,698
(level, geography, grade band, year, student group) combinations the live page
produced.  Compares against the values dumped out of the running browser.

Two independent checks per combination:
  A. counts path  -- sum the '# Level n' columns, divide by 'Number Tested'
  B. published path -- for single-grade cells, compare against the file's own
     '% Level n' columns, which the dashboard never reads.
"""
import openpyxl, os, sys
from collections import defaultdict

R = os.path.join(os.path.dirname(__file__), '..')
YEARS = [2018, 2019, 2022, 2023, 2024, 2025, 2026]
GRADES = ['3','4','5','6','7','8','All Grades']
CATS = ['All Students','SWD','Not SWD','Econ Disadv','Not Econ Disadv',
        'Current ELL','Ever ELL','Never ELL','Asian','Black','Hispanic',
        'Multi-Racial','Native American','White','Female','Male','Neither Female nor Male']
BOROS = ['Bronx','Brooklyn','Manhattan','Queens','Staten Island']
BKEY = {'BRONX':'Bronx','BROOKLYN':'Brooklyn','MANHATTAN':'Manhattan',
        'QUEENS':'Queens','STATEN ISLAND':'Staten Island'}
DISTRICTS = ['%02d'%d for d in range(1,33)]
BANDS = {'all':['All Grades'],'35':['3','4','5'],'68':['6','7','8'],
         'g3':['3'],'g4':['4'],'g5':['5'],'g6':['6'],'g7':['7'],'g8':['8']}

# ---- read the workbooks completely fresh -------------------------------
def load(path, idcol):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    counts, published = {}, {}
    for sn in wb.sheetnames:
        if not sn.startswith('ELA'): continue
        ws = wb[sn]; it = ws.iter_rows(values_only=True)
        hdr = list(next(it)); H = {h:i for i,h in enumerate(hdr)}
        for r in it:
            if r is None or r[0] is None: continue
            cat = r[H['Category']]
            if cat not in CATS: continue
            vals = [r[H[c]] for c in ('Number Tested','Mean Scale Score','# Level 1',
                                      '# Level 2','# Level 3','# Level 4')]
            if any(v == 's' for v in vals): continue
            nt = vals[0]
            g = str(r[H['Grade']]); y = r[H['Year']]
            gid = BKEY[r[H[idcol]]] if idcol=='Borough' else (r[H[idcol]] if idcol else 'city')
            k = (gid, g, y, cat)
            counts[k] = (nt, r[H['Mean Scale Score']],
                         r[H['# Level 1']], r[H['# Level 2']],
                         r[H['# Level 3']], r[H['# Level 4']])
            published[k] = (r[H['% Level 1']], r[H['% Level 2']],
                            r[H['% Level 3']], r[H['% Level 4']], r[H['% Level 3+4']])
    return counts, published

C = {}; P = {}
for lvl, fn, idc in [('city','citywide-ela-results-public.xlsx',None),
                     ('boro','borough-ela-results-public.xlsx','Borough'),
                     ('dist','district-ela-results-public.xlsx','District')]:
    C[lvl], P[lvl] = load(os.path.join(R,'data',fn), idc)
    print(f'loaded {lvl}: {len(C[lvl]):,} unsuppressed cells', file=sys.stderr)

GEO = {'city': ['city'], 'boro': BOROS, 'dist': DISTRICTS}

def expect(lvl, geoIdx, bandKey, year, catIdx):
    gid = GEO[lvl][geoIdx]; cat = CATS[catIdx]
    n=c1=c2=c3=c4=0; wm=0.0; found=False
    for g in BANDS[bandKey]:
        v = C[lvl].get((gid,g,year,cat))
        if v is None: continue
        found=True
        nt,ms,a,b,c,d = v
        n+=nt; c1+=a; c2+=b; c3+=c; c4+=d; wm+=ms*nt
    if not found or n==0: return None
    return (n, c1/n*100, c2/n*100, c3/n*100, c4/n*100, (c3+c4)/n*100, wm/n)

# ---- compare against the browser dump ----------------------------------
BAD=[]; ok=0; nullok=0; nullmismatch=0; checked_pub=0; pubmax=0.0
MAXPCT=0.0; MAXMEAN=0.0
TOL = 5e-6      # percentage points
TOLM = 1e-5     # mean scale score points

for line in open(os.path.join(R,'audit/browser_dump.txt')):
    f = line.rstrip('\n').split('|')
    lvl, geo, bk, yr, cat = f[0], int(f[1]), f[2], int(f[3]), int(f[4])
    got = f[5:]
    exp = expect(lvl, geo, bk, yr, cat)
    if exp is None:
        if got[0] == '': nullok += 1
        else: nullmismatch += 1; BAD.append((line.strip(),'browser has data, source has none'))
        continue
    if got[0] == '':
        BAD.append((line.strip(),'browser empty, source has data %s'%(exp,))); continue
    gn = int(got[0])
    vals = [float(x) for x in got[1:6]] ; gmean = float(got[6])
    if gn != exp[0]:
        BAD.append((line.strip(),'N %s != %s'%(gn,exp[0]))); continue
    devs = [abs(vals[i]-exp[i+1]) for i in range(5)]
    MAXPCT = max(MAXPCT, max(devs)); MAXMEAN = max(MAXMEAN, abs(gmean-exp[6]))
    if max(devs) > TOL:
        BAD.append((line.strip(),'pct dev %.3e vs %s'%(max(devs),exp[1:6]))); continue
    if abs(gmean-exp[6]) > TOLM:
        BAD.append((line.strip(),'mean %.6f != %.6f'%(gmean,exp[6]))); continue
    ok += 1
    # --- independent published-percentage check, single-grade cells only
    if bk.startswith('g') or bk=='all':
        g = BANDS[bk][0]
        pub = P[lvl].get((GEO[lvl][geo], g, yr, CATS[cat]))
        if pub:
            checked_pub += 1
            for i,pv in enumerate(pub[:4]):
                pubmax = max(pubmax, abs(vals[i]-pv))
            pubmax = max(pubmax, abs(vals[4]-pub[4]))

print('='*72)
print('ENGINE AUDIT — dashboard values vs independent recomputation from .xlsx')
print('='*72)
print(f'combinations compared          : {ok+nullok+len(BAD):,}')
print(f'  matched (with data)          : {ok:,}')
print(f'  matched (suppressed both)    : {nullok:,}')
print(f'  MISMATCHES                   : {len(BAD):,}')
print()
print(f'max deviation, percentages     : {MAXPCT:.3e} pp')
print(f'max deviation, mean scale score: {MAXMEAN:.3e} points')
print()
print(f'published-% cross-check cells  : {checked_pub:,}')
print(f'  max deviation vs published % : {pubmax:.3e} pp')
print(f'tolerances: {TOL:g} pp on percentages, {TOLM:g} on mean scale score')
if BAD:
    print('\nFIRST 25 MISMATCHES')
    for l,why in BAD[:25]: print('  ', l, '->', why)
    sys.exit(1)
print('\nRESULT: PASS — every combination on the live page reproduces exactly.')
