import openpyxl, collections, math, sys
BASE='data/%s-ela-results-public.xlsx'
LEVELS=['sheet','rows','suppressed_s','other_nonnum','sum_counts_ne_tested','pct_mismatch>0.01','pct_sum_ne_100','l34_ne_l3+l4','pctl34_mismatch']

def analyze(level):
    wb=openpyxl.load_workbook(BASE%level, read_only=True, data_only=True)
    out=[]
    for sn in wb.sheetnames:
        if not sn.startswith('ELA'): continue
        ws=wb[sn]
        rows=list(ws.iter_rows(values_only=True))
        hdr=list(rows[0]); rows=rows[1:]
        H={h:i for i,h in enumerate(hdr)}
        n=len(rows); s_count=0; other=0; sum_ne=0; pmis=0; psum=0; l34ne=0; p34mis=0
        maxdev_pct=0; maxdev_p34=0; maxdev_psum=0
        for r in rows:
            vals=[r[H[c]] for c in ['Number Tested','Mean Scale Score','# Level 1','# Level 2','# Level 3','# Level 4','# Level 3+4']]
            pcts=[r[H[c]] for c in ['% Level 1','% Level 2','% Level 3','% Level 4','% Level 3+4']]
            allv=vals+pcts
            if any(v=='s' for v in allv): s_count+=1; continue
            bad=[v for v in allv if not isinstance(v,(int,float))]
            if bad: other+=1; continue
            nt,ms,c1,c2,c3,c4,c34=vals
            p1,p2,p3,p4,p34=pcts
            if c1+c2+c3+c4!=nt: sum_ne+=1
            if c3+c4!=c34: l34ne+=1
            if nt:
                for c,p in zip([c1,c2,c3,c4],[p1,p2,p3,p4]):
                    d=abs(c/nt*100-p); maxdev_pct=max(maxdev_pct,d)
                    if d>0.01: pmis+=1
                d=abs(c34/nt*100-p34); maxdev_p34=max(maxdev_p34,d)
                if d>0.01: p34mis+=1
            d=abs(p1+p2+p3+p4-100); maxdev_psum=max(maxdev_psum,d)
            if d>0.01: psum+=1
        out.append((level,sn,n,s_count,other,sum_ne,pmis,psum,l34ne,p34mis,round(maxdev_pct,6),round(maxdev_p34,6),round(maxdev_psum,6)))
    return out

print(f"{'lvl':<9}{'sheet':<26}{'rows':>6}{'sup_s':>7}{'oth':>5}{'cnt!=nt':>8}{'pct!=':>6}{'sum!=100':>9}{'l34!=':>6}{'p34!=':>6}{'mxdPct':>9}{'mxdP34':>9}{'mxdSum':>9}")
for lvl in ['citywide','borough','district']:
    for row in analyze(lvl):
        print(f"{row[0]:<9}{row[1]:<26}{row[2]:>6}{row[3]:>7}{row[4]:>5}{row[5]:>8}{row[6]:>6}{row[7]:>9}{row[8]:>6}{row[9]:>6}{row[10]:>9}{row[11]:>9}{row[12]:>9}")
