"""Engine audit by digest, over exact integers.

Recomputes all 40,698 (level, geography, grade band, year, student group)
combinations directly from the original .xlsx and hashes the aggregated student
COUNTS -- number tested and the four level counts. Those are integers, so the
comparison is exact and carries no dependence on decimal formatting.

Counts are the right thing to hash. Every percentage the dashboard shows is
count / tested computed at render time, so identical counts imply identical
percentages by construction. An earlier version of this script hashed the
formatted percentages instead and reported spurious failures: 548/2048 is
exactly 26.7578125, and Python rounds that tie down to 26.757812 while
JavaScript's toFixed rounds it up to 26.757813. The values were identical; only
the tie-break convention differed. Mean scale score, the one genuinely
floating-point field, is checked numerically with a tolerance in
audit_engine.py rather than by digest.
"""
import openpyxl, os, hashlib, sys
R=os.path.join(os.path.dirname(__file__),'..')
YEARS=[2018,2019,2022,2023,2024,2025,2026]
CATS=['All Students','SWD','Not SWD','Econ Disadv','Not Econ Disadv','Current ELL','Ever ELL',
      'Never ELL','Asian','Black','Hispanic','Multi-Racial','Native American','White',
      'Female','Male','Neither Female nor Male']
BOROS=['Bronx','Brooklyn','Manhattan','Queens','Staten Island']
BKEY={'BRONX':'Bronx','BROOKLYN':'Brooklyn','MANHATTAN':'Manhattan','QUEENS':'Queens','STATEN ISLAND':'Staten Island'}
DISTRICTS=['%02d'%d for d in range(1,33)]
BAND_ORDER=['all','35','68','g3','g4','g5','g6','g7','g8']
BANDS={'all':['All Grades'],'35':['3','4','5'],'68':['6','7','8'],
       'g3':['3'],'g4':['4'],'g5':['5'],'g6':['6'],'g7':['7'],'g8':['8']}

def load(path, idcol):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True); C={}
    for sn in wb.sheetnames:
        if not sn.startswith('ELA'): continue
        ws=wb[sn]; it=ws.iter_rows(values_only=True); hdr=list(next(it)); H={h:i for i,h in enumerate(hdr)}
        for r in it:
            if r is None or r[0] is None: continue
            cat=r[H['Category']]
            if cat not in CATS: continue
            v=[r[H[c]] for c in ('Number Tested','Mean Scale Score','# Level 1','# Level 2','# Level 3','# Level 4')]
            if any(x=='s' for x in v): continue
            # mirror the payload exactly: the mean is stored to 6 decimals, so
            # the audit must weight the same value the page weights. Every other
            # field is an integer count and is carried through unchanged.
            v[1] = round(float(v[1]), 6)
            gid=BKEY[r[H[idcol]]] if idcol=='Borough' else (r[H[idcol]] if idcol else 'city')
            C[(gid,str(r[H['Grade']]),r[H['Year']],cat)]=v
    return C
C={'city':load(os.path.join(R,'data','citywide-ela-results-public.xlsx'),None),
   'boro':load(os.path.join(R,'data','borough-ela-results-public.xlsx'),'Borough'),
   'dist':load(os.path.join(R,'data','district-ela-results-public.xlsx'),'District')}
GEO={'city':['city'],'boro':BOROS,'dist':DISTRICTS}

def row(lvl, gi, bk, y, ci):
    gid=GEO[lvl][gi]; cat=CATS[ci]
    n=c1=c2=c3=c4=0; wm=0.0; found=False
    for g in BANDS[bk]:
        v=C[lvl].get((gid,g,y,cat))
        if v is None: continue
        found=True; nt,ms,a,b,c,d=v
        n+=nt;c1+=a;c2+=b;c3+=c;c4+=d;wm+=ms*nt
    if not found or n==0:
        return '|'.join([lvl,str(gi),bk,str(y),str(ci),'','','','',''])
    return '|'.join([lvl,str(gi),bk,str(y),str(ci),
                     str(n),str(c1),str(c2),str(c3),str(c4)])

rows=[]
for bk in BAND_ORDER:
    for y in YEARS:
        for ci in range(len(CATS)): rows.append(row('city',0,bk,y,ci))
for gi in range(len(BOROS)):
    for bk in BAND_ORDER:
        for y in YEARS:
            for ci in range(len(CATS)): rows.append(row('boro',gi,bk,y,ci))
for gi in range(len(DISTRICTS)):
    for bk in BAND_ORDER:
        for y in YEARS:
            for ci in range(len(CATS)): rows.append(row('dist',gi,bk,y,ci))
text='\n'.join(rows)
h=hashlib.sha256(text.encode()).hexdigest()
BROWSER = sys.argv[1] if len(sys.argv)>1 else None
print('='*72)
print('ENGINE AUDIT (digest) — aggregated student counts, all combinations')
print('='*72)
print(f'combinations : {len(rows):,}')
print(f'chars        : {len(text):,}')
print(f'sha256       : {h}')
if BROWSER:
    print(f'live page    : {BROWSER}')
    if h==BROWSER:
        print('\nRESULT: PASS — counts identical to the running dashboard for all')
        print('40,698 combinations, so every derived percentage is identical too.')
    else:
        print('\nRESULT: FAIL — the dashboard does not match the source files.')
        sys.exit(1)
